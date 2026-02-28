"""
Excel export and review module for NEET Test Generator.
Handles generation of type-specific Excel files (MCQ, AR, MTC)
and reading/updating Excel files for the review workflow.
"""

import io
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ============================================================
# LATEX → UNICODE CONVERSION
# ============================================================

_SUP = {'0': '\u2070', '1': '\u00b9', '2': '\u00b2', '3': '\u00b3', '4': '\u2074',
        '5': '\u2075', '6': '\u2076', '7': '\u2077', '8': '\u2078', '9': '\u2079',
        '+': '\u207a', '-': '\u207b', 'n': '\u207f', '(': '\u207d', ')': '\u207e'}

_SUB = {'0': '\u2080', '1': '\u2081', '2': '\u2082', '3': '\u2083', '4': '\u2084',
        '5': '\u2085', '6': '\u2086', '7': '\u2087', '8': '\u2088', '9': '\u2089',
        '+': '\u208a', '-': '\u208b', '(': '\u208d', ')': '\u208e',
        'a': '\u2090', 'e': '\u2091', 'o': '\u2092', 'x': '\u2093',
        'h': '\u2095', 'k': '\u2096', 'l': '\u2097', 'm': '\u2098',
        'n': '\u2099', 'p': '\u209a', 's': '\u209b', 't': '\u209c'}

_GREEK = {
    'alpha': '\u03b1', 'beta': '\u03b2', 'gamma': '\u03b3', 'delta': '\u03b4',
    'pi': '\u03c0', 'sigma': '\u03c3', 'lambda': '\u03bb', 'mu': '\u03bc',
    'theta': '\u03b8', 'Delta': '\u0394',
}


def _to_superscript(s: str) -> str:
    return ''.join(_SUP.get(c, c) for c in s)


def _to_subscript(s: str) -> str:
    return ''.join(_SUB.get(c, c) for c in s)


def latex_to_unicode(text: str) -> str:
    """Convert LaTeX math notation to clean Unicode text.
    Handles: $...$, ^{}, _{}, \\alpha, \\Delta H, \\cdot, \\text{}, etc.
    """
    if not text:
        return ""
    text = str(text)

    # Strip $ delimiters
    text = re.sub(r'\$([^$]+)\$', r'\1', text)

    # Remove \text{} and \mathrm{} wrappers
    text = re.sub(r'\\(?:text|mathrm|textrm)\{([^}]*)\}', r'\1', text)

    # Greek letters: \alpha → α
    for name, char in _GREEK.items():
        text = text.replace(f'\\{name}', char)

    # \cdot → ·
    text = text.replace('\\cdot', '\u00b7')

    # ^\circ → ° (degree, often after E)
    text = re.sub(r'\^\\circ', '\u00b0', text)
    # standalone \circ → °
    text = text.replace('\\circ', '\u00b0')

    # \rightarrow / \to → →
    text = text.replace('\\rightarrow', '\u2192')
    text = text.replace('\\to', '\u2192')

    # \times → ×
    text = text.replace('\\times', '\u00d7')

    # ^{...} → superscript
    text = re.sub(r'\^{([^}]*)}', lambda m: _to_superscript(m.group(1)), text)

    # ^X (single char) → superscript
    text = re.sub(r'\^([0-9+\-n])', lambda m: _to_superscript(m.group(1)), text)

    # _{...} → subscript
    text = re.sub(r'_{([^}]*)}', lambda m: _to_subscript(m.group(1)), text)

    # _X (single char) → subscript (digits and common letters)
    text = re.sub(r'_([0-9aehklmnopstx])', lambda m: _to_subscript(m.group(1)), text)

    # Clean leftover backslashes from unknown commands
    text = re.sub(r'\\([a-zA-Z]+)', r'\1', text)

    return text


# ============================================================
# UTILITIES
# ============================================================

def _clean(text: str) -> str:
    """Clean text for Excel: convert LaTeX to Unicode, normalize whitespace."""
    text = latex_to_unicode(str(text) if text else "")
    text = text.replace('\\n\\n', '\n').replace('\\n', '\n')
    return text.strip()


def _format_time(seconds) -> str:
    """Format seconds as 'Xm Ys' or 'Xs'."""
    if seconds is None or seconds == "":
        return ""
    try:
        s = float(seconds)
    except (ValueError, TypeError):
        return str(seconds)
    if s >= 60:
        return f"{int(s // 60)}m {int(s % 60)}s"
    return f"{s:.1f}s"


# ============================================================
# STYLING
# ============================================================

_HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
_HEADER_FILL = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
_HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
_CELL_ALIGN = Alignment(vertical='top', wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def _style_headers(ws, num_cols: int):
    """Apply header styling to row 1."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _auto_width(ws):
    """Set reasonable column widths."""
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = 0
        for cell in col:
            if cell.value:
                lines = str(cell.value).split('\n')
                max_len = max(max_len, max(len(line) for line in lines))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)


def _add_meta_row(ws, row: int, col_start: int, metadata: dict):
    """Add Time, token, and cost columns. Populated on EVERY row."""
    token_usage = metadata.get("token_usage", {})
    gen = token_usage.get("generation", {}) or {}
    cost = token_usage.get("cost", {}) or {}

    ws.cell(row=row, column=col_start).value = _format_time(metadata.get("generation_time", ""))
    ws.cell(row=row, column=col_start + 1).value = gen.get("input_tokens", "")
    ws.cell(row=row, column=col_start + 2).value = gen.get("output_tokens", "")
    ws.cell(row=row, column=col_start + 3).value = cost.get("input_cost", "")
    ws.cell(row=row, column=col_start + 4).value = cost.get("output_cost", "")


# ============================================================
# AR PARSING
# ============================================================

def _parse_ar(question_text: str):
    """Extract Assertion and Reason from AR question_text."""
    text = str(question_text) if question_text else ""
    # Try "Assertion (A):" / "Reason (R):" pattern
    m = re.search(r'Assertion\s*\(?A\)?\s*:\s*(.*?)\s*Reason\s*\(?R\)?\s*:', text, re.DOTALL | re.IGNORECASE)
    assertion = m.group(1).strip() if m else text
    r = re.search(r'Reason\s*\(?R\)?\s*:\s*(.*)', text, re.DOTALL | re.IGNORECASE)
    reason = r.group(1).strip() if r else ""
    return _clean(assertion), _clean(reason)


# ============================================================
# MTC PARSING
# ============================================================

def _parse_mtc(question_text: str):
    """Extract List I items and List II items from MTC question_text."""
    text = str(question_text) if question_text else ""
    list_i_items = []
    list_ii_items = []

    # Try pipe-table format: "A. xxx | I. yyy"
    pairs = re.findall(r'([A-D])\.\s*(.*?)\s*\|\s*(IV|III|II|I)\.\s*(.*?)(?=\s+[A-D]\.\s|\s*Choose|\s*$)', text)
    if pairs:
        for letter, l1, roman, l2 in pairs:
            list_i_items.append(f"{letter}. {_clean(l1)}")
            list_ii_items.append(f"{roman}. {_clean(l2)}")
        return "\n".join(list_i_items), "\n".join(list_ii_items)

    # Fallback: return full text as List I
    return _clean(text), ""


# ============================================================
# SHEET BUILDERS
# ============================================================

MCQ_HEADERS = [
    "Question Number", "Question", "Option A", "Option B", "Option C", "Option D",
    "Correct Answer", "Accuracy", "Comment",
    "Time to Run", "Input Tokens", "Output Tokens",
    "Input Cost (Rs)", "Output Cost (Rs)"
]

AR_HEADERS = [
    "Question Number", "Assertion (A)", "Reason (R)", "Correct Answer",
    "Accuracy", "Comment",
    "Time to Run", "Input Tokens", "Output Tokens",
    "Input Cost (Rs)", "Output Cost (Rs)"
]

MTC_HEADERS = [
    "Question Number", "List I", "List II", "Correct Answer",
    "Accuracy", "Comment",
    "Time to Run", "Input Tokens", "Output Tokens",
    "Input Cost (Rs)", "Output Cost (Rs)"
]


def _build_mcq_sheet(ws, questions: list, metadata: dict):
    for col, h in enumerate(MCQ_HEADERS, 1):
        ws.cell(row=1, column=col, value=h)
    _style_headers(ws, len(MCQ_HEADERS))

    for idx, q in enumerate(questions):
        row = idx + 2
        opts = q.get("options", {})
        ws.cell(row=row, column=1).value = q.get("question_id", idx + 1)
        ws.cell(row=row, column=2).value = _clean(q.get("question_text", ""))
        ws.cell(row=row, column=3).value = _clean(opts.get("a", ""))
        ws.cell(row=row, column=4).value = _clean(opts.get("b", ""))
        ws.cell(row=row, column=5).value = _clean(opts.get("c", ""))
        ws.cell(row=row, column=6).value = _clean(opts.get("d", ""))
        ws.cell(row=row, column=7).value = (q.get("correct_answer") or "").upper()
        # 8=Accuracy, 9=Comment left blank
        _add_meta_row(ws, row, 10, metadata)

        for c in range(1, len(MCQ_HEADERS) + 1):
            ws.cell(row=row, column=c).alignment = _CELL_ALIGN

    _auto_width(ws)


def _build_ar_sheet(ws, questions: list, metadata: dict):
    for col, h in enumerate(AR_HEADERS, 1):
        ws.cell(row=1, column=col, value=h)
    _style_headers(ws, len(AR_HEADERS))

    for idx, q in enumerate(questions):
        row = idx + 2
        assertion, reason = _parse_ar(q.get("question_text", ""))
        ws.cell(row=row, column=1).value = q.get("question_id", idx + 1)
        ws.cell(row=row, column=2).value = assertion
        ws.cell(row=row, column=3).value = reason
        ws.cell(row=row, column=4).value = (q.get("correct_answer") or "").upper()
        # 5=Accuracy, 6=Comment left blank
        _add_meta_row(ws, row, 7, metadata)

        for c in range(1, len(AR_HEADERS) + 1):
            ws.cell(row=row, column=c).alignment = _CELL_ALIGN

    _auto_width(ws)


def _build_mtc_sheet(ws, questions: list, metadata: dict):
    for col, h in enumerate(MTC_HEADERS, 1):
        ws.cell(row=1, column=col, value=h)
    _style_headers(ws, len(MTC_HEADERS))

    for idx, q in enumerate(questions):
        row = idx + 2
        list_i, list_ii = _parse_mtc(q.get("question_text", ""))
        ws.cell(row=row, column=1).value = q.get("question_id", idx + 1)
        ws.cell(row=row, column=2).value = list_i
        ws.cell(row=row, column=3).value = list_ii
        ws.cell(row=row, column=4).value = (q.get("correct_answer") or "").upper()
        # 5=Accuracy, 6=Comment left blank
        _add_meta_row(ws, row, 7, metadata)

        for c in range(1, len(MTC_HEADERS) + 1):
            ws.cell(row=row, column=c).alignment = _CELL_ALIGN

    _auto_width(ws)


# ============================================================
# PUBLIC API — GENERATE
# ============================================================

def generate_excel_for_result(result: dict) -> bytes:
    """Generate an Excel file from a generation result dict. Returns bytes."""
    metadata = result.get("test_metadata", {})
    questions = result.get("questions", [])
    question_type = metadata.get("question_type", "mcq")

    wb = Workbook()
    ws = wb.active

    if question_type == "combination":
        mcq_qs = [q for q in questions if q.get("question_type", "").upper() == "MCQ"]
        ar_qs = [q for q in questions if q.get("question_type", "").upper() in ("ASSERTION_REASON", "ASSERTION-REASON", "AR")]
        mtc_qs = [q for q in questions if q.get("question_type", "").upper() in ("MATCH_THE_COLUMN", "MTC")]

        sheets_created = False
        if mcq_qs:
            ws.title = "MCQ"
            _build_mcq_sheet(ws, mcq_qs, metadata)
            sheets_created = True
        if ar_qs:
            ar_ws = wb.create_sheet("Assertion-Reason") if sheets_created else ws
            if not sheets_created:
                ar_ws.title = "Assertion-Reason"
                sheets_created = True
            _build_ar_sheet(ar_ws, ar_qs, metadata)
        if mtc_qs:
            mtc_ws = wb.create_sheet("Match the Column") if sheets_created else ws
            if not sheets_created:
                mtc_ws.title = "Match the Column"
                sheets_created = True
            _build_mtc_sheet(mtc_ws, mtc_qs, metadata)

        if not sheets_created:
            ws.title = "MCQ"
            _build_mcq_sheet(ws, questions, metadata)
    elif question_type in ("assertion_reason",):
        ws.title = "Assertion-Reason"
        _build_ar_sheet(ws, questions, metadata)
    elif question_type in ("match_the_column",):
        ws.title = "Match the Column"
        _build_mtc_sheet(ws, questions, metadata)
    else:
        ws.title = "MCQ"
        _build_mcq_sheet(ws, questions, metadata)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# PUBLIC API — REVIEW
# ============================================================

def read_excel_for_review(excel_bytes: bytes) -> tuple:
    """Read an exported Excel file for review.
    Returns (list_of_question_dicts, question_type_str).
    Handles multi-sheet (combination) workbooks.
    """
    wb = load_workbook(io.BytesIO(excel_bytes))
    all_questions = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1] if cell.value]
        if not headers:
            continue

        # Detect type from headers
        if "Assertion (A)" in headers:
            qtype = "ASSERTION_REASON"
        elif "List I" in headers:
            qtype = "MATCH_THE_COLUMN"
        else:
            qtype = "MCQ"

        header_map = {h: i for i, h in enumerate(headers)}

        for row in ws.iter_rows(min_row=2, max_col=len(headers)):
            vals = [cell.value for cell in row]
            if not vals[0]:  # skip empty rows
                continue

            q = {"_type": qtype, "_sheet": sheet_name, "_row": row[0].row}

            if qtype == "MCQ":
                q["question_text"] = vals[header_map.get("Question", 1)] or ""
                q["options"] = {
                    "a": vals[header_map.get("Option A", 2)] or "",
                    "b": vals[header_map.get("Option B", 3)] or "",
                    "c": vals[header_map.get("Option C", 4)] or "",
                    "d": vals[header_map.get("Option D", 5)] or "",
                }
                q["correct_answer"] = vals[header_map.get("Correct Answer", 6)] or ""
                q["accuracy"] = vals[header_map.get("Accuracy", 7)] or ""
                q["comment"] = vals[header_map.get("Comment", 8)] or ""
            elif qtype == "ASSERTION_REASON":
                q["assertion"] = vals[header_map.get("Assertion (A)", 1)] or ""
                q["reason"] = vals[header_map.get("Reason (R)", 2)] or ""
                q["correct_answer"] = vals[header_map.get("Correct Answer", 3)] or ""
                q["accuracy"] = vals[header_map.get("Accuracy", 4)] or ""
                q["comment"] = vals[header_map.get("Comment", 5)] or ""
            elif qtype == "MATCH_THE_COLUMN":
                q["list_i"] = vals[header_map.get("List I", 1)] or ""
                q["list_ii"] = vals[header_map.get("List II", 2)] or ""
                q["correct_answer"] = vals[header_map.get("Correct Answer", 3)] or ""
                q["accuracy"] = vals[header_map.get("Accuracy", 4)] or ""
                q["comment"] = vals[header_map.get("Comment", 5)] or ""

            all_questions.append(q)

    # Determine overall type
    types = {q["_type"] for q in all_questions}
    if len(types) == 1:
        overall_type = types.pop()
    else:
        overall_type = "COMBINATION"

    return all_questions, overall_type


def update_excel_with_comments(excel_bytes: bytes, comments: dict, accuracies: dict = None) -> bytes:
    """Write comments (and optionally accuracy) back into the Excel file.
    comments = {question_index: comment_text}
    accuracies = {question_index: accuracy_text} (optional)
    """
    wb = load_workbook(io.BytesIO(excel_bytes))

    # Build a mapping: question_index -> (sheet, excel_row)
    q_idx = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1] if cell.value]
        if not headers:
            continue

        comment_col = None
        accuracy_col = None
        for i, h in enumerate(headers):
            if h == "Comment":
                comment_col = i + 1
            if h == "Accuracy":
                accuracy_col = i + 1

        for row_num in range(2, ws.max_row + 1):
            if ws.cell(row=row_num, column=1).value is None:
                continue
            if comment_col and q_idx in comments:
                ws.cell(row=row_num, column=comment_col).value = comments[q_idx]
            if accuracy_col and accuracies and q_idx in accuracies:
                ws.cell(row=row_num, column=accuracy_col).value = accuracies[q_idx]
            q_idx += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
